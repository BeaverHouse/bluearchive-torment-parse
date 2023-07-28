from openpyxl import load_workbook
from openpyxl.cell import Cell
import json

book = load_workbook("data.xlsx", data_only=True)

# 빈 셀 확인하는 함수
def is_cell_blank(cell: Cell) -> bool:
    return cell.value is None or len(str(cell.value).strip()) <= 0

# 파티가 존재하는지 확인하는 함수
# 스트라이커만 있어도 유효 파티
def party_exists(row, i:int) -> bool:
    if i*7 + 9 > len(row): return False
    return not (
        is_cell_blank(row[i*7+3]) \
        and is_cell_blank(row[i*7+4]) \
        and is_cell_blank(row[i*7+5]) \
        and is_cell_blank(row[i*7+6])
    )


for idx, name in enumerate(book.sheetnames):

    offset = 1 if idx>=2 else 0

    total_json = []

    sheet = book[name]

    for row in sheet.iter_rows(min_row=2+offset):
        dic = {}

        if is_cell_blank(row[offset]): break

        # 점수 순위 저장
        dic["rank"] = int(row[offset].value)
        dic["score"] = int(row[1+offset].value)

        i = 0
        partys = []

        # 전체 캐릭명 저장할 임시 array
        temp_char_array = []
        
        # 파티가 존재하면 파티 추가
        while party_exists(row, i):
            party = {}
            strikers = list(filter(
                lambda x1: x1 is not None and len(x1) > 0,
                list(map(lambda x2: x2.value, row[i*7+3:i*7+7]))
            ))
            specials = list(filter(
                lambda x1: x1 is not None and len(x1) > 0,
                list(map(lambda x2: x2.value, row[i*7+7:i*7+9]))
            ))
            party["strikers"] = strikers
            party["specials"] = specials

            # 임시 array에 캐릭터들 추가
            temp_char_array += strikers + specials

            partys.append(party)
            
            i += 1

        dic["partys"] = partys
        dic["party_count"] = len(partys)

        # 캐릭터 사용현황 dict에 정리
        total_char_dic = {}
        for char in temp_char_array:
            total_char_dic[char] = total_char_dic.get(char, 0) + 1
        
        dic["characters"] = list(total_char_dic.keys())
        
        # 2번쓰면 무조건 조력자
        for char in total_char_dic.keys():
            if total_char_dic.get(char, 0) == 2:
                dic["assist"] = char
                break

        print(".", end="")
        total_json.append(dic)

    # JSON 저장
    with open("{}.json".format(name.split()[0]), "w", encoding='utf-8') as f:
        json.dump(total_json, f, indent=2, ensure_ascii=False)



